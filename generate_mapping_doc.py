"""
Generate the mapping document Excel file for usp_LoadDimCustomer.
This script is for generation only and should NOT be committed to the repo.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Styles
HEADER_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
SUBHEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
LABEL_FONT = Font(name="Calibri", size=11, bold=True)
LABEL_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
NORMAL_FONT = Font(name="Calibri", size=11)
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
TRANSFORM_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
WRAP_ALIGN = Alignment(wrap_text=True, vertical="top")
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)


def style_header_row(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER


def style_data_rows(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER
            cell.alignment = WRAP_ALIGN


def auto_width(ws, max_col, min_width=12, max_width=50):
    for col in range(1, max_col + 1):
        max_len = min_width
        for row in ws.iter_rows(min_col=col, max_col=col, values_only=False):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, min(len(str(cell.value)), max_width))
        ws.column_dimensions[get_column_letter(col)].width = max_len + 2


wb = openpyxl.Workbook()

# ============================================================================
# TAB 1: Header
# ============================================================================
ws = wb.active
ws.title = "Header"

header_data = [
    ("Document Title", "Mapping Document - usp_LoadDimCustomer"),
    ("Project Name", "R_POC - Data Warehouse Proof of Concept"),
    ("Module", "DimCustomer Dimension Load"),
    ("Stored Procedure", "dbo.usp_LoadDimCustomer"),
    ("Source System", "Upstream CRM (via Staging_Customer)"),
    ("Target System", "DW_POC SQL Server Database"),
    ("Source Table", "dbo.Staging_Customer"),
    ("Target Table", "dbo.DimCustomer"),
    ("Load Strategy", "MERGE - SCD Type 1 (Overwrite)"),
    ("Author", "Devin"),
    ("Created Date", "2026-03-27"),
    ("Version", "1.0"),
    ("Status", "Draft"),
    ("Description", "Transforms staged customer data and loads it into the DimCustomer dimension table using MERGE (SCD Type 1). Applies 5 transformations: FullName concatenation, Age calculation, IncomeCategory bucketing, CustomerSegment derivation, and IsActive flag."),
    ("Dependencies", "Requires Staging_Customer to be populated by usp_ExtractToStaging before execution."),
    ("Execution Frequency", "Daily / On-demand"),
    ("Estimated Row Volume", "10,000 - 1,000,000 rows"),
    ("Error Handling", "TRY/CATCH with ETL_ExecutionLog logging"),
]

ws.column_dimensions["A"].width = 25
ws.column_dimensions["B"].width = 80

for i, (label, value) in enumerate(header_data, start=1):
    cell_a = ws.cell(row=i, column=1, value=label)
    cell_a.font = LABEL_FONT
    cell_a.fill = LABEL_FILL
    cell_a.border = THIN_BORDER
    cell_a.alignment = WRAP_ALIGN

    cell_b = ws.cell(row=i, column=2, value=value)
    cell_b.font = NORMAL_FONT
    cell_b.border = THIN_BORDER
    cell_b.alignment = WRAP_ALIGN

# ============================================================================
# TAB 2: Data Set
# ============================================================================
ws2 = wb.create_sheet("Data Set")

dataset_headers = ["Data Set Name", "Database", "Schema", "Table/View", "Type", "Row Count (Est.)", "Description"]
for c, h in enumerate(dataset_headers, 1):
    ws2.cell(row=1, column=c, value=h)
style_header_row(ws2, 1, len(dataset_headers))

datasets = [
    ("Source - Staging Customer", "DW_POC", "dbo", "Staging_Customer", "Staging Table", "Variable", "Intermediate landing zone populated by usp_ExtractToStaging from upstream CRM system"),
    ("Target - DimCustomer", "DW_POC", "dbo", "DimCustomer", "Dimension Table", "Variable", "Customer dimension table with 25 columns including 5 derived/transformed columns"),
    ("Logging - ETL Execution Log", "DW_POC", "dbo", "ETL_ExecutionLog", "Log Table", "N/A", "Audit/execution log capturing procedure start, success, failure, and row counts"),
    ("Temp - Merge Output", "tempdb", "dbo", "#MergeOutput", "Temp Table", "Variable", "Temporary table used to capture MERGE OUTPUT action counts (INSERT/UPDATE)"),
]

for r, row_data in enumerate(datasets, 2):
    for c, val in enumerate(row_data, 1):
        ws2.cell(row=r, column=c, value=val)

style_data_rows(ws2, 2, len(datasets) + 1, len(dataset_headers))
auto_width(ws2, len(dataset_headers))

# ============================================================================
# TAB 3: Source Data
# ============================================================================
ws3 = wb.create_sheet("Source Data")

source_headers = ["#", "Column Name", "Data Type", "Nullable", "Primary Key", "Description", "Sample Value"]
for c, h in enumerate(source_headers, 1):
    ws3.cell(row=1, column=c, value=h)
style_header_row(ws3, 1, len(source_headers))

source_cols = [
    (1, "CustomerID", "NVARCHAR(20)", "No", "Yes", "Unique business key from upstream CRM", "CUST-001"),
    (2, "FirstName", "NVARCHAR(100)", "Yes", "No", "Customer first name", "John"),
    (3, "LastName", "NVARCHAR(100)", "Yes", "No", "Customer last name", "Smith"),
    (4, "Email", "NVARCHAR(256)", "Yes", "No", "Customer email address", "john.smith@example.com"),
    (5, "Phone", "NVARCHAR(50)", "Yes", "No", "Customer phone number", "555-0101"),
    (6, "DateOfBirth", "DATE", "Yes", "No", "Customer date of birth", "1985-03-15"),
    (7, "Gender", "NVARCHAR(10)", "Yes", "No", "Customer gender", "Male"),
    (8, "MaritalStatus", "NVARCHAR(20)", "Yes", "No", "Customer marital status", "Married"),
    (9, "AddressLine1", "NVARCHAR(256)", "Yes", "No", "Primary address line", "123 Main St"),
    (10, "AddressLine2", "NVARCHAR(256)", "Yes", "No", "Secondary address line (apt, suite, etc.)", "Apt 4B"),
    (11, "City", "NVARCHAR(100)", "Yes", "No", "City name", "New York"),
    (12, "StateProvince", "NVARCHAR(100)", "Yes", "No", "State or province", "NY"),
    (13, "PostalCode", "NVARCHAR(20)", "Yes", "No", "Postal/ZIP code", "10001"),
    (14, "Country", "NVARCHAR(100)", "Yes", "No", "Country name", "USA"),
    (15, "AnnualIncome", "DECIMAL(18,2)", "Yes", "No", "Customer annual income", "85000.00"),
    (16, "TotalPurchaseAmount", "DECIMAL(18,2)", "Yes", "No", "Lifetime total purchase amount", "12500.00"),
    (17, "AccountOpenDate", "DATE", "Yes", "No", "Date account was opened", "2018-01-10"),
    (18, "AccountStatus", "NVARCHAR(20)", "Yes", "No", "Current account status (Active, Inactive, Closed, Suspended)", "Active"),
    (19, "PreferredContactMethod", "NVARCHAR(50)", "Yes", "No", "Preferred contact method (Email, Phone, Mail, SMS)", "Email"),
    (20, "LastModifiedDate", "DATETIME", "Yes", "No", "Last modification timestamp from source", "2026-03-27 00:00:00"),
]

for r, row_data in enumerate(source_cols, 2):
    for c, val in enumerate(row_data, 1):
        ws3.cell(row=r, column=c, value=val)

style_data_rows(ws3, 2, len(source_cols) + 1, len(source_headers))
auto_width(ws3, len(source_headers))

# ============================================================================
# TAB 4: Target Data
# ============================================================================
ws4 = wb.create_sheet("Target Data")

target_headers = ["#", "Column Name", "Data Type", "Nullable", "Primary Key", "Unique", "Default", "Column Type", "Description", "Sample Value"]
for c, h in enumerate(target_headers, 1):
    ws4.cell(row=1, column=c, value=h)
style_header_row(ws4, 1, len(target_headers))

target_cols = [
    (1, "CustomerKey", "INT IDENTITY(1,1)", "No", "Yes", "No", "Auto-increment", "Surrogate Key", "Auto-generated surrogate key", "1"),
    (2, "CustomerID", "NVARCHAR(20)", "No", "No", "Yes", "-", "Business Key", "Business key from source system", "CUST-001"),
    (3, "FirstName", "NVARCHAR(100)", "Yes", "No", "No", "-", "1:1 Mapping", "Customer first name", "John"),
    (4, "LastName", "NVARCHAR(100)", "Yes", "No", "No", "-", "1:1 Mapping", "Customer last name", "Smith"),
    (5, "FullName", "NVARCHAR(201)", "Yes", "No", "No", "-", "Transformed", "Full name (FirstName + LastName)", "John Smith"),
    (6, "Email", "NVARCHAR(256)", "Yes", "No", "No", "-", "1:1 Mapping", "Customer email address", "john.smith@example.com"),
    (7, "Phone", "NVARCHAR(50)", "Yes", "No", "No", "-", "1:1 Mapping", "Customer phone number", "555-0101"),
    (8, "DateOfBirth", "DATE", "Yes", "No", "No", "-", "1:1 Mapping", "Customer date of birth", "1985-03-15"),
    (9, "Age", "INT", "Yes", "No", "No", "-", "Transformed", "Calculated age from DateOfBirth", "41"),
    (10, "Gender", "NVARCHAR(10)", "Yes", "No", "No", "-", "1:1 Mapping", "Customer gender", "Male"),
    (11, "MaritalStatus", "NVARCHAR(20)", "Yes", "No", "No", "-", "1:1 Mapping", "Customer marital status", "Married"),
    (12, "AddressLine1", "NVARCHAR(256)", "Yes", "No", "No", "-", "1:1 Mapping", "Primary address line", "123 Main St"),
    (13, "AddressLine2", "NVARCHAR(256)", "Yes", "No", "No", "-", "1:1 Mapping", "Secondary address line", "Apt 4B"),
    (14, "City", "NVARCHAR(100)", "Yes", "No", "No", "-", "1:1 Mapping", "City name", "New York"),
    (15, "StateProvince", "NVARCHAR(100)", "Yes", "No", "No", "-", "1:1 Mapping", "State or province", "NY"),
    (16, "PostalCode", "NVARCHAR(20)", "Yes", "No", "No", "-", "1:1 Mapping", "Postal/ZIP code", "10001"),
    (17, "Country", "NVARCHAR(100)", "Yes", "No", "No", "-", "1:1 Mapping", "Country name", "USA"),
    (18, "AnnualIncome", "DECIMAL(18,2)", "Yes", "No", "No", "-", "1:1 Mapping", "Customer annual income", "85000.00"),
    (19, "IncomeCategory", "NVARCHAR(20)", "Yes", "No", "No", "-", "Transformed", "Income bucket: Low/Medium/High/Very High", "High"),
    (20, "TotalPurchaseAmount", "DECIMAL(18,2)", "Yes", "No", "No", "-", "1:1 Mapping", "Lifetime total purchase amount", "12500.00"),
    (21, "CustomerSegment", "NVARCHAR(20)", "Yes", "No", "No", "-", "Transformed", "Segment: Bronze/Silver/Gold/Platinum", "Gold"),
    (22, "AccountOpenDate", "DATE", "Yes", "No", "No", "-", "1:1 Mapping", "Date account was opened", "2018-01-10"),
    (23, "AccountStatus", "NVARCHAR(20)", "Yes", "No", "No", "-", "1:1 Mapping", "Current account status", "Active"),
    (24, "IsActive", "BIT", "No", "No", "No", "0", "Transformed", "Active flag: 1 if Active, else 0", "1"),
    (25, "LoadDate", "DATETIME", "No", "No", "No", "GETDATE()", "System", "ETL load timestamp", "2026-03-27 02:15:00"),
]

for r, row_data in enumerate(target_cols, 2):
    for c, val in enumerate(row_data, 1):
        ws4.cell(row=r, column=c, value=val)
    # Highlight transformed columns
    col_type = row_data[7]
    if col_type == "Transformed":
        for c in range(1, len(target_headers) + 1):
            ws4.cell(row=r, column=c).fill = TRANSFORM_FILL

style_data_rows(ws4, 2, len(target_cols) + 1, len(target_headers))
auto_width(ws4, len(target_headers))

# ============================================================================
# TAB 5: Data Dictionary
# ============================================================================
ws5 = wb.create_sheet("Data Dictionary")

dict_headers = ["#", "Column Name", "Business Name", "Data Type", "Length/Precision", "Nullable", "Domain / Valid Values", "Business Definition", "Source System", "Notes"]
for c, h in enumerate(dict_headers, 1):
    ws5.cell(row=1, column=c, value=h)
style_header_row(ws5, 1, len(dict_headers))

dict_data = [
    (1, "CustomerKey", "Customer Key", "INT", "4 bytes", "No", "Auto-increment integers", "System-generated surrogate key uniquely identifying each customer record in the dimension", "System Generated", "IDENTITY column; do not insert manually"),
    (2, "CustomerID", "Customer Identifier", "NVARCHAR", "20 chars", "No", "Format: CUST-NNN", "Unique business identifier assigned by the upstream CRM system", "Upstream CRM", "Business key used for MERGE matching"),
    (3, "FirstName", "First Name", "NVARCHAR", "100 chars", "Yes", "Free text", "Customer's given/first name", "Upstream CRM", "Direct 1:1 mapping"),
    (4, "LastName", "Last Name", "NVARCHAR", "100 chars", "Yes", "Free text", "Customer's family/last name", "Upstream CRM", "Direct 1:1 mapping"),
    (5, "FullName", "Full Name", "NVARCHAR", "201 chars", "Yes", "Free text", "Concatenation of FirstName and LastName with space separator", "Derived", "Transformation: LTRIM(RTRIM(COALESCE(FirstName,'') + ' ' + COALESCE(LastName,'')))"),
    (6, "Email", "Email Address", "NVARCHAR", "256 chars", "Yes", "Valid email format", "Customer's primary email address", "Upstream CRM", "Direct 1:1 mapping"),
    (7, "Phone", "Phone Number", "NVARCHAR", "50 chars", "Yes", "Free text", "Customer's primary phone number", "Upstream CRM", "Direct 1:1 mapping"),
    (8, "DateOfBirth", "Date of Birth", "DATE", "3 bytes", "Yes", "Valid past dates", "Customer's date of birth", "Upstream CRM", "Direct 1:1 mapping; used to derive Age"),
    (9, "Age", "Customer Age", "INT", "4 bytes", "Yes", "0-150 or NULL", "Customer's current age in years, calculated from DateOfBirth", "Derived", "Transformation: DATEDIFF with birthday adjustment; NULL if DOB is NULL or future"),
    (10, "Gender", "Gender", "NVARCHAR", "10 chars", "Yes", "Male, Female, Other, NULL", "Customer's gender identity", "Upstream CRM", "Direct 1:1 mapping"),
    (11, "MaritalStatus", "Marital Status", "NVARCHAR", "20 chars", "Yes", "Single, Married, Divorced, Widowed, NULL", "Customer's marital status", "Upstream CRM", "Direct 1:1 mapping"),
    (12, "AddressLine1", "Address Line 1", "NVARCHAR", "256 chars", "Yes", "Free text", "Primary street address", "Upstream CRM", "Direct 1:1 mapping"),
    (13, "AddressLine2", "Address Line 2", "NVARCHAR", "256 chars", "Yes", "Free text", "Secondary address (apartment, suite, unit)", "Upstream CRM", "Direct 1:1 mapping"),
    (14, "City", "City", "NVARCHAR", "100 chars", "Yes", "Free text", "City of residence", "Upstream CRM", "Direct 1:1 mapping"),
    (15, "StateProvince", "State / Province", "NVARCHAR", "100 chars", "Yes", "Free text / state codes", "State or province of residence", "Upstream CRM", "Direct 1:1 mapping"),
    (16, "PostalCode", "Postal Code", "NVARCHAR", "20 chars", "Yes", "Free text / ZIP format", "Postal or ZIP code", "Upstream CRM", "Direct 1:1 mapping"),
    (17, "Country", "Country", "NVARCHAR", "100 chars", "Yes", "Free text / country names", "Country of residence", "Upstream CRM", "Direct 1:1 mapping"),
    (18, "AnnualIncome", "Annual Income", "DECIMAL", "18,2", "Yes", ">= 0 or NULL", "Customer's reported annual income in USD", "Upstream CRM", "Direct 1:1 mapping; used to derive IncomeCategory"),
    (19, "IncomeCategory", "Income Category", "NVARCHAR", "20 chars", "Yes", "Low, Medium, High, Very High, Unknown", "Categorization of annual income into predefined buckets", "Derived", "Transformation: CASE on AnnualIncome thresholds"),
    (20, "TotalPurchaseAmount", "Total Purchase Amount", "DECIMAL", "18,2", "Yes", ">= 0 or NULL", "Lifetime cumulative purchase amount in USD", "Upstream CRM", "Direct 1:1 mapping; used to derive CustomerSegment"),
    (21, "CustomerSegment", "Customer Segment", "NVARCHAR", "20 chars", "Yes", "Bronze, Silver, Gold, Platinum, Unknown", "Customer tier based on total purchase amount", "Derived", "Transformation: CASE on TotalPurchaseAmount thresholds"),
    (22, "AccountOpenDate", "Account Open Date", "DATE", "3 bytes", "Yes", "Valid past dates", "Date when the customer account was first opened", "Upstream CRM", "Direct 1:1 mapping"),
    (23, "AccountStatus", "Account Status", "NVARCHAR", "20 chars", "Yes", "Active, Inactive, Closed, Suspended", "Current status of the customer account", "Upstream CRM", "Direct 1:1 mapping; used to derive IsActive"),
    (24, "IsActive", "Is Active Flag", "BIT", "1 byte", "No", "0 or 1", "Boolean flag indicating whether customer account is currently active", "Derived", "Transformation: 1 if UPPER(AccountStatus)='ACTIVE', else 0"),
    (25, "LoadDate", "Load Date", "DATETIME", "8 bytes", "No", "Valid datetime", "Timestamp when the record was inserted or last updated by ETL", "System Generated", "Set to GETDATE() on INSERT and UPDATE"),
]

for r, row_data in enumerate(dict_data, 2):
    for c, val in enumerate(row_data, 1):
        ws5.cell(row=r, column=c, value=val)

style_data_rows(ws5, 2, len(dict_data) + 1, len(dict_headers))
auto_width(ws5, len(dict_headers))

# ============================================================================
# TAB 6: Mapping
# ============================================================================
ws6 = wb.create_sheet("Mapping")

map_headers = ["#", "Source Column", "Source Data Type", "Target Column", "Target Data Type", "Mapping Type", "Mapping Rule / Logic", "Join Key", "NULL Handling", "Notes"]
for c, h in enumerate(map_headers, 1):
    ws6.cell(row=1, column=c, value=h)
style_header_row(ws6, 1, len(map_headers))

map_data = [
    (1, "-", "-", "CustomerKey", "INT IDENTITY", "System Generated", "Auto-increment surrogate key", "-", "NOT NULL (auto)", "Do not map; generated by SQL Server"),
    (2, "CustomerID", "NVARCHAR(20)", "CustomerID", "NVARCHAR(20)", "1:1 Direct", "Direct copy", "Yes (MERGE ON)", "NOT NULL", "Business key used for MERGE match"),
    (3, "FirstName", "NVARCHAR(100)", "FirstName", "NVARCHAR(100)", "1:1 Direct", "Direct copy", "No", "Allow NULL", ""),
    (4, "LastName", "NVARCHAR(100)", "LastName", "NVARCHAR(100)", "1:1 Direct", "Direct copy", "No", "Allow NULL", ""),
    (5, "FirstName, LastName", "NVARCHAR(100) each", "FullName", "NVARCHAR(201)", "Transformation", "LTRIM(RTRIM(COALESCE(FirstName,'') + CASE WHEN both NOT NULL THEN ' ' ELSE '' END + COALESCE(LastName,'')))", "No", "Empty string if both NULL", "Concatenation with NULL-safe handling"),
    (6, "Email", "NVARCHAR(256)", "Email", "NVARCHAR(256)", "1:1 Direct", "Direct copy", "No", "Allow NULL", ""),
    (7, "Phone", "NVARCHAR(50)", "Phone", "NVARCHAR(50)", "1:1 Direct", "Direct copy", "No", "Allow NULL", ""),
    (8, "DateOfBirth", "DATE", "DateOfBirth", "DATE", "1:1 Direct", "Direct copy", "No", "Allow NULL", ""),
    (9, "DateOfBirth", "DATE", "Age", "INT", "Transformation", "DATEDIFF(YEAR, DateOfBirth, GETDATE()) - birthday_adjustment", "No", "NULL if DOB is NULL or future date", "Recalculated on each load; may drift between runs"),
    (10, "Gender", "NVARCHAR(10)", "Gender", "NVARCHAR(10)", "1:1 Direct", "Direct copy", "No", "Allow NULL", ""),
    (11, "MaritalStatus", "NVARCHAR(20)", "MaritalStatus", "NVARCHAR(20)", "1:1 Direct", "Direct copy", "No", "Allow NULL", ""),
    (12, "AddressLine1", "NVARCHAR(256)", "AddressLine1", "NVARCHAR(256)", "1:1 Direct", "Direct copy", "No", "Allow NULL", ""),
    (13, "AddressLine2", "NVARCHAR(256)", "AddressLine2", "NVARCHAR(256)", "1:1 Direct", "Direct copy", "No", "Allow NULL", ""),
    (14, "City", "NVARCHAR(100)", "City", "NVARCHAR(100)", "1:1 Direct", "Direct copy", "No", "Allow NULL", ""),
    (15, "StateProvince", "NVARCHAR(100)", "StateProvince", "NVARCHAR(100)", "1:1 Direct", "Direct copy", "No", "Allow NULL", ""),
    (16, "PostalCode", "NVARCHAR(20)", "PostalCode", "NVARCHAR(20)", "1:1 Direct", "Direct copy", "No", "Allow NULL", ""),
    (17, "Country", "NVARCHAR(100)", "Country", "NVARCHAR(100)", "1:1 Direct", "Direct copy", "No", "Allow NULL", ""),
    (18, "AnnualIncome", "DECIMAL(18,2)", "AnnualIncome", "DECIMAL(18,2)", "1:1 Direct", "Direct copy", "No", "Allow NULL", ""),
    (19, "AnnualIncome", "DECIMAL(18,2)", "IncomeCategory", "NVARCHAR(20)", "Transformation", "CASE: NULL->'Unknown', <30K->'Low', <75K->'Medium', <150K->'High', else->'Very High'", "No", "'Unknown' if NULL", "Bucketed categorization"),
    (20, "TotalPurchaseAmount", "DECIMAL(18,2)", "TotalPurchaseAmount", "DECIMAL(18,2)", "1:1 Direct", "Direct copy", "No", "Allow NULL", ""),
    (21, "TotalPurchaseAmount", "DECIMAL(18,2)", "CustomerSegment", "NVARCHAR(20)", "Transformation", "CASE: NULL->'Unknown', <1K->'Bronze', <10K->'Silver', <50K->'Gold', else->'Platinum'", "No", "'Unknown' if NULL", "Tier-based segmentation"),
    (22, "AccountOpenDate", "DATE", "AccountOpenDate", "DATE", "1:1 Direct", "Direct copy", "No", "Allow NULL", ""),
    (23, "AccountStatus", "NVARCHAR(20)", "AccountStatus", "NVARCHAR(20)", "1:1 Direct", "Direct copy", "No", "Allow NULL", ""),
    (24, "AccountStatus", "NVARCHAR(20)", "IsActive", "BIT", "Transformation", "CASE WHEN UPPER(AccountStatus) = 'ACTIVE' THEN 1 ELSE 0 END", "No", "Defaults to 0", "Case-insensitive comparison"),
    (25, "-", "-", "LoadDate", "DATETIME", "System Generated", "GETDATE() at time of INSERT or UPDATE", "-", "NOT NULL", "ETL metadata timestamp"),
]

for r, row_data in enumerate(map_data, 2):
    for c, val in enumerate(row_data, 1):
        ws6.cell(row=r, column=c, value=val)
    mapping_type = row_data[5]
    if mapping_type == "Transformation":
        for c in range(1, len(map_headers) + 1):
            ws6.cell(row=r, column=c).fill = TRANSFORM_FILL

style_data_rows(ws6, 2, len(map_data) + 1, len(map_headers))
auto_width(ws6, len(map_headers))

# ============================================================================
# TAB 7: Health Checks
# ============================================================================
ws7 = wb.create_sheet("Health Checks")

hc_headers = ["#", "Check Category", "Check Name", "Check Description", "SQL Query / Validation", "Expected Result", "Severity", "Frequency"]
for c, h in enumerate(hc_headers, 1):
    ws7.cell(row=1, column=c, value=h)
style_header_row(ws7, 1, len(hc_headers))

hc_data = [
    (1, "Row Count", "Source vs Target Count", "Verify row count in staging matches DimCustomer after initial load", "SELECT COUNT(*) FROM Staging_Customer;\nSELECT COUNT(*) FROM DimCustomer;", "Counts should match (initial load) or target >= source", "Critical", "Every Run"),
    (2, "Duplicate Check", "Duplicate CustomerID", "Ensure no duplicate business keys in DimCustomer", "SELECT CustomerID, COUNT(*) FROM DimCustomer GROUP BY CustomerID HAVING COUNT(*) > 1;", "Zero rows returned", "Critical", "Every Run"),
    (3, "NULL Check", "NULL Business Keys", "Ensure CustomerID is never NULL in target", "SELECT COUNT(*) FROM DimCustomer WHERE CustomerID IS NULL;", "Zero", "Critical", "Every Run"),
    (4, "Transformation", "FullName Validation", "Verify FullName matches FirstName + LastName concatenation", "SELECT * FROM DimCustomer WHERE FullName <> LTRIM(RTRIM(COALESCE(FirstName,'') + ' ' + COALESCE(LastName,'')));", "Zero rows returned", "High", "Every Run"),
    (5, "Transformation", "Age Validation", "Verify Age is correctly calculated from DateOfBirth", "SELECT * FROM DimCustomer WHERE DateOfBirth IS NOT NULL AND Age IS NULL;", "Zero rows returned", "High", "Every Run"),
    (6, "Transformation", "IncomeCategory Validation", "Verify IncomeCategory aligns with AnnualIncome thresholds", "SELECT * FROM DimCustomer WHERE AnnualIncome < 30000 AND IncomeCategory <> 'Low' AND AnnualIncome IS NOT NULL;", "Zero rows returned", "High", "Every Run"),
    (7, "Transformation", "CustomerSegment Validation", "Verify CustomerSegment aligns with TotalPurchaseAmount thresholds", "SELECT * FROM DimCustomer WHERE TotalPurchaseAmount >= 50000 AND CustomerSegment <> 'Platinum';", "Zero rows returned", "High", "Every Run"),
    (8, "Transformation", "IsActive Flag Validation", "Verify IsActive matches AccountStatus", "SELECT * FROM DimCustomer WHERE AccountStatus = 'Active' AND IsActive = 0;", "Zero rows returned", "High", "Every Run"),
    (9, "Referential", "Orphan Records Check", "Check for records in target not present in staging (unexpected deletes scenario)", "SELECT d.CustomerID FROM DimCustomer d LEFT JOIN Staging_Customer s ON d.CustomerID = s.CustomerID WHERE s.CustomerID IS NULL;", "Review any orphaned records", "Medium", "Weekly"),
    (10, "Data Quality", "Email Format Check", "Validate email addresses contain @ symbol", "SELECT COUNT(*) FROM DimCustomer WHERE Email IS NOT NULL AND Email NOT LIKE '%@%.%';", "Zero", "Low", "Weekly"),
    (11, "Data Quality", "Future DateOfBirth", "Ensure no future birth dates exist in target", "SELECT COUNT(*) FROM DimCustomer WHERE DateOfBirth > GETDATE();", "Zero", "Medium", "Every Run"),
    (12, "Data Quality", "Negative Income", "Ensure no negative annual income values", "SELECT COUNT(*) FROM DimCustomer WHERE AnnualIncome < 0;", "Zero", "Medium", "Every Run"),
    (13, "ETL Audit", "Execution Log Check", "Verify ETL execution logged successfully", "SELECT TOP 1 * FROM ETL_ExecutionLog WHERE ProcedureName = 'usp_LoadDimCustomer' ORDER BY LogID DESC;", "Status = 'Success'", "Critical", "Every Run"),
    (14, "ETL Audit", "LoadDate Currency", "Verify LoadDate is current for recently loaded records", "SELECT COUNT(*) FROM DimCustomer WHERE LoadDate < DATEADD(DAY, -1, GETDATE());", "Review stale records", "Medium", "Daily"),
    (15, "Performance", "Merge Duration", "Monitor ETL execution time", "SELECT DATEDIFF(SECOND, StartTime, EndTime) FROM ETL_ExecutionLog WHERE ProcedureName = 'usp_LoadDimCustomer' AND Status = 'Success' ORDER BY LogID DESC;", "Within acceptable SLA", "Medium", "Every Run"),
]

for r, row_data in enumerate(hc_data, 2):
    for c, val in enumerate(row_data, 1):
        ws7.cell(row=r, column=c, value=val)

style_data_rows(ws7, 2, len(hc_data) + 1, len(hc_headers))
auto_width(ws7, len(hc_headers))

# ============================================================================
# TAB 8: Transformation
# ============================================================================
ws8 = wb.create_sheet("Transformation")

tf_headers = ["#", "Transformation Name", "Target Column", "Source Column(s)", "Transformation Type", "Business Rule", "SQL Logic", "NULL Handling", "Edge Cases", "Example Input", "Example Output"]
for c, h in enumerate(tf_headers, 1):
    ws8.cell(row=1, column=c, value=h)
style_header_row(ws8, 1, len(tf_headers))

tf_data = [
    (
        1,
        "Full Name Concatenation",
        "FullName",
        "FirstName, LastName",
        "String Concatenation",
        "Combine first and last name into a single full name field with a space separator. Handle NULL values gracefully so that if either name is missing, no leading/trailing spaces appear.",
        "LTRIM(RTRIM(\n  COALESCE(s.FirstName, N'') +\n  CASE WHEN s.FirstName IS NOT NULL\n       AND s.LastName IS NOT NULL\n       THEN N' ' ELSE N'' END +\n  COALESCE(s.LastName, N'')\n))",
        "COALESCE to empty string; no space if either is NULL",
        "Both NULL -> empty string; One NULL -> only the non-NULL name",
        "FirstName='John', LastName='Smith'",
        "'John Smith'",
    ),
    (
        2,
        "Age Calculation",
        "Age",
        "DateOfBirth",
        "Date Arithmetic",
        "Calculate the customer's current age in whole years from their date of birth. Adjust for whether the birthday has occurred this year. Return NULL for NULL or future birth dates.",
        "CASE\n  WHEN s.DateOfBirth IS NULL THEN NULL\n  WHEN s.DateOfBirth > GETDATE() THEN NULL\n  ELSE DATEDIFF(YEAR, s.DateOfBirth, GETDATE())\n    - CASE WHEN DATEADD(YEAR,\n        DATEDIFF(YEAR, s.DateOfBirth, GETDATE()),\n        s.DateOfBirth) > GETDATE()\n      THEN 1 ELSE 0 END\nEND",
        "NULL DateOfBirth -> NULL Age",
        "Future DOB -> NULL; Leap year birthdays handled by DATEADD",
        "DateOfBirth='1985-03-15' (run date 2026-03-27)",
        "41",
    ),
    (
        3,
        "Income Categorization",
        "IncomeCategory",
        "AnnualIncome",
        "Bucketed Classification",
        "Classify customers into income tiers based on their annual income:\n- NULL -> 'Unknown'\n- < $30,000 -> 'Low'\n- $30,000 to $74,999 -> 'Medium'\n- $75,000 to $149,999 -> 'High'\n- >= $150,000 -> 'Very High'",
        "CASE\n  WHEN s.AnnualIncome IS NULL THEN N'Unknown'\n  WHEN s.AnnualIncome < 30000 THEN N'Low'\n  WHEN s.AnnualIncome < 75000 THEN N'Medium'\n  WHEN s.AnnualIncome < 150000 THEN N'High'\n  ELSE N'Very High'\nEND",
        "NULL -> 'Unknown'",
        "Negative income -> 'Low'; Zero income -> 'Low'; Boundary values (e.g., 30000 exactly) -> 'Medium'",
        "AnnualIncome=85000.00",
        "'High'",
    ),
    (
        4,
        "Customer Segmentation",
        "CustomerSegment",
        "TotalPurchaseAmount",
        "Bucketed Classification",
        "Segment customers by their lifetime purchase value:\n- NULL -> 'Unknown'\n- < $1,000 -> 'Bronze'\n- $1,000 to $9,999 -> 'Silver'\n- $10,000 to $49,999 -> 'Gold'\n- >= $50,000 -> 'Platinum'",
        "CASE\n  WHEN s.TotalPurchaseAmount IS NULL THEN N'Unknown'\n  WHEN s.TotalPurchaseAmount < 1000 THEN N'Bronze'\n  WHEN s.TotalPurchaseAmount < 10000 THEN N'Silver'\n  WHEN s.TotalPurchaseAmount < 50000 THEN N'Gold'\n  ELSE N'Platinum'\nEND",
        "NULL -> 'Unknown'",
        "Negative amounts -> 'Bronze'; Zero -> 'Bronze'; Boundary (1000 exactly) -> 'Silver'",
        "TotalPurchaseAmount=12500.00",
        "'Gold'",
    ),
    (
        5,
        "Active Flag Derivation",
        "IsActive",
        "AccountStatus",
        "Boolean Conversion",
        "Convert the text-based AccountStatus into a boolean flag. Only 'Active' (case-insensitive) maps to 1 (true). All other statuses (Inactive, Closed, Suspended, NULL) map to 0 (false).",
        "CASE\n  WHEN UPPER(s.AccountStatus) = N'ACTIVE'\n  THEN 1\n  ELSE 0\nEND",
        "NULL AccountStatus -> 0 (not active)",
        "Case variations ('active', 'ACTIVE', 'Active') all map to 1",
        "AccountStatus='Active'",
        "1",
    ),
]

for r, row_data in enumerate(tf_data, 2):
    for c, val in enumerate(row_data, 1):
        ws8.cell(row=r, column=c, value=val)

style_data_rows(ws8, 2, len(tf_data) + 1, len(tf_headers))
auto_width(ws8, len(tf_headers))

# ============================================================================
# Save
# ============================================================================
output_path = "/home/ubuntu/R_POC/docs/MappingDocument_usp_LoadDimCustomer.xlsx"
import os
os.makedirs(os.path.dirname(output_path), exist_ok=True)
wb.save(output_path)
print(f"Mapping document saved to: {output_path}")
